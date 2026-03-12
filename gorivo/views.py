import openpyxl
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.utils.dateparse import parse_datetime
from django.db.models import Q
from .models import Tank, TransakcijaGoriva
from .forms import PotrosnjaForm, RefillForm, RaspodjelaForm
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


def gorivo_view(request):
    tank = Tank.objects.first()
    if not tank:
        tank = Tank.objects.create()

    potrosnja_form = PotrosnjaForm()
    refill_form = RefillForm()
    raspodjela_form = RaspodjelaForm(tank=tank)

    # GET parametri za filter
    firma = request.GET.get("firma")  # "bih" ili "rh"
    datum = request.GET.get("datum")  # datetime string
    vozilo_id = request.GET.get("vozilo")  # id vozila

    transakcije = tank.transakcije.all()

    filter_applied = False

    # filtriranje po firmi
    if firma in ["bih", "rh"]:
        transakcije = transakcije.filter(drzava=firma)
        filter_applied = True

    # filtriranje po datumu
    if datum:
        dt = parse_datetime(datum)
        if dt:
            transakcije = transakcije.filter(datum__gte=dt)
            filter_applied = True

    # filtriranje po vozilu
    if vozilo_id:
        transakcije = transakcije.filter(vozilo_id=vozilo_id)
        filter_applied = True

    # default prikaz – zadnjih 20 samo ako filter nije aktivan
    if not filter_applied:
        transakcije = transakcije.order_by("-datum")[:20]
    else:
        transakcije = transakcije.order_by("-datum")  # sve filtrirane, bez limitiranja
        
    # Dohvati sva vozila koja imaju transakcije u ovom tanku
    vozila = (
        tank.transakcije
        .filter(vozilo__isnull=False)
        .values_list("vozilo__id", "vozilo__ime")
        .distinct()
    )

    context = {
        "tank": tank,
        "potrosnja_form": potrosnja_form,
        "refill_form": refill_form,
        "form_raspodjela": raspodjela_form,
        "transakcije": transakcije,
        "filter_firma": firma,
        "filter_datum": datum,
        "filter_vozilo": vozilo_id,
        "vozila": vozila,
    }

    return render(request, "gorivo/gorivo.html", context)


def dodaj_potrosnju(request):
    tank = Tank.objects.first()
    if not tank:
        tank = Tank.objects.create()

    if request.method == "POST":
        form = PotrosnjaForm(request.POST)
        if form.is_valid():
            transakcija = form.save(commit=False)
            transakcija.tank = tank
            transakcija.tip = "potrosnja"
            transakcija.drzava = request.POST.get("drzava")  # bih ili rh
            transakcija.save()
    return redirect("gorivo")

def dodaj_potrosnju(request):
    tank = Tank.objects.first()
    if not tank:
        tank = Tank.objects.create()

    if request.method == "POST":
        form = PotrosnjaForm(request.POST)
        if form.is_valid():
            try:
                transakcija = form.save(commit=False)
                transakcija.tank = tank
                transakcija.tip = "potrosnja"
                transakcija.drzava = request.POST.get("drzava")  # bih ili rh
                transakcija.save()
                messages.success(request, "Potrošnja uspješno spremljena.")
            except ValidationError as e:
                # direktno u messages.error
                if hasattr(e, "message_dict"):
                    poruke = []
                    for polje, msgs in e.message_dict.items():
                        poruke.extend(msgs)
                    messages.error(request, " ".join(poruke))
                else:
                    messages.error(request, str(e))

            return redirect("gorivo")

        else:
            messages.error(request, "Molimo popunite sve obavezne podatke ispravno.")
            return redirect("gorivo")

    return redirect("gorivo")


def dodaj_refill(request):
    tank = Tank.objects.first()

    if not tank:
        tank = Tank.objects.create()

    if request.method == "POST":
        form = RefillForm(request.POST, instance=tank)

        if form.is_valid():

            bih = form.cleaned_data.get("bih_gorivo") or 0
            rh = form.cleaned_data.get("rh_gorivo") or 0

            ukupno_za_dodati = bih + rh
            trenutno = tank.ukupno_stanje

            # PROVJERA KAPACITETA
            if trenutno + ukupno_za_dodati > tank.kapacitet:

                slobodno = tank.kapacitet - trenutno

                messages.error(
                    request,
                    f"Nema dovoljno mjesta u tanku. "
                    f"Slobodno je još {slobodno} L."
                )

                return redirect("gorivo")

            # dodaj refill BIH
            if bih > 0:
                TransakcijaGoriva.objects.create(
                    tank=tank,
                    tip="refill",
                    drzava="bih",
                    kolicina=bih,
                    napomena="Punjenje tanka BiH"
                )

            # dodaj refill RH
            if rh > 0:
                TransakcijaGoriva.objects.create(
                    tank=tank,
                    tip="refill",
                    drzava="rh",
                    kolicina=rh,
                    napomena="Punjenje tanka RH"
                )

            # update kapaciteta ako je promijenjen
            tank.kapacitet = form.cleaned_data.get("kapacitet", tank.kapacitet)
            tank.save()

            messages.success(request, "Punjenje tanka uspješno spremljeno.")

    return redirect("gorivo")

def raspodjela_view(request):
    tank = Tank.objects.first()
    if not tank:
        tank = Tank.objects.create()

    if request.method == "POST":
        form = RaspodjelaForm(request.POST, tank=tank)
        if form.is_valid():
            # Prvo resetiramo sve raspodjele goriva
            tank.transakcije.filter(tip="raspodjela").delete()

            # Dohvatimo unose
            bih = form.cleaned_data["bih_gorivo"]
            rh = form.cleaned_data["rh_gorivo"]

            # Ako je uneseno > 0, kreiramo transakcije sa tip="raspodjela"
            if bih > 0:
                TransakcijaGoriva.objects.create(
                    tank=tank,
                    tip="potrosnja",  # Ovdje koristimo potrosnja za oduzimanje
                    drzava="bih",
                    kolicina=bih,
                    napomena="Potrošeno gorivo BiH"
                )
            if rh > 0:
                TransakcijaGoriva.objects.create(
                    tank=tank,
                    tip="potrosnja",
                    drzava="rh",
                    kolicina=rh,
                    napomena="Potrošeno gorivo RH"
                )

            # Ukupni kapacitet samo update ako se promijeni
            tank.kapacitet = form.cleaned_data["kapacitet"]
            tank.save()
            return redirect("gorivo")
    else:
        form = RaspodjelaForm(tank=tank)

    return render(request, "gorivo/raspodjela.html", {"form": form})


def export_excel(request):
    tank = Tank.objects.first()
    if not tank:
        tank = Tank.objects.create()

    # GET parametri za filter
    firma = request.GET.get("firma")
    datum = request.GET.get("datum")
    vozilo_id = request.GET.get("vozilo")

    transakcije = tank.transakcije.all().order_by("datum")  # sortiramo po datumu

    # filtriranje
    if firma in ["bih", "rh"]:
        transakcije = transakcije.filter(drzava=firma)
    if datum:
        dt = parse_datetime(datum)
        if dt:
            transakcije = transakcije.filter(datum__gte=dt)
    if vozilo_id:
        transakcije = transakcije.filter(vozilo_id=vozilo_id)

    # Kreiraj Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transakcije goriva"

    # Zaglavlja
    headers = ["Datum", "Vozilo", "Ulaz goriva (L)", "Izlaz goriva (L)", "Stanje (L)", "Napomena"]
    ws.append(headers)

    # Pratimo stanje po firmi
    stanje_firme = {
        "bih": 0,
        "rh": 0,
    }

    # Izračunaj početno stanje prije prve transakcije
    for t in tank.transakcije.filter(drzava="bih"):
        if t.tip == "refill":
            stanje_firme["bih"] += t.kolicina
        elif t.tip == "potrosnja":
            stanje_firme["bih"] -= t.kolicina

    for t in tank.transakcije.filter(drzava="rh"):
        if t.tip == "refill":
            stanje_firme["rh"] += t.kolicina
        elif t.tip == "potrosnja":
            stanje_firme["rh"] -= t.kolicina

    # za filtrirane transakcije, pišemo red po red
    # i pratimo stanje firme
    stanje_firme = {
        "bih": 0,
        "rh": 0,
    }

    for t in transakcije:
        drzava = t.drzava
        ulaz = stanje_firme.get(drzava, 0)
        if t.tip == "refill":
            stanje_firme[drzava] = ulaz + t.kolicina
            izlaz = 0
        else:  # potrosnja
            stanje_firme[drzava] = ulaz
            izlaz = t.kolicina
            ulaz = ulaz  # stanje prije potrosnje

        ukupno_stanje = tank.stanje_bih + tank.stanje_rh

        ws.append([
            t.datum.strftime("%d.%m.%Y %H:%M"),
            str(t.vozilo) if t.vozilo else "-",
            round(ulaz,2) if t.tip=="refill" else 0,
            round(izlaz,2) if t.tip=="potrosnja" else 0,
            round(tank.ukupno_stanje,2),
            t.napomena or "",
        ])

        # update stanje firme
        if t.tip == "refill":
            stanje_firme[drzava] += t.kolicina
        elif t.tip == "potrosnja":
            stanje_firme[drzava] -= t.kolicina

    # podesimo širinu kolona
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # pripremi HttpResponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=transakcije_goriva.xlsx'
    wb.save(response)
    return response