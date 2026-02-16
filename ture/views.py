import os
from datetime import datetime, timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.conf import settings
from reportlab.lib.pagesizes import A4, landscape # type: ignore
from reportlab.pdfbase import pdfmetrics# type: ignore
from reportlab.pdfbase.ttfonts import TTFont# type: ignore
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, KeepInFrame # type: ignore
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle# type: ignore
from reportlab.lib import colors# type: ignore
from .models import Tura, Vozac, Vozilo, Naputak, RadniNalog, osvjezi_radni_nalog, CijenaDnevnica
from .forms import TuraForm, VozacForm, VozacUpdateForm, VoziloForm, NaputakForm, RadniNalogForm
from openpyxl import load_workbook
from config.settings import BASE_DIR
from pathlib import Path

def parse_int(value, default=None):
    if value in (None, '', 'None'):
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default

def login_view(request):
    if request.user.is_authenticated:
        return redirect('homepage') 

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('homepage')
        else:
            messages.error(request, 'Neispravno korisničko ime ili lozinka.')

    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def homepage(request):
    mjesec = parse_int(request.GET.get('mjesec'))
    godina = parse_int(request.GET.get('godina'))
    status = request.GET.get('status', 'aktivne')  # po defaultu aktivne
    # Odredi aktivan status prema odabiru
    aktivan_filter = True if status == 'aktivne' else False

    if godina is not None:
        if mjesec is not None:
            ture = Tura.objects.filter(
                aktivan=aktivan_filter,
                datum_polaska__month=mjesec,
                datum_polaska__year=godina
            )
        else:
            ture = Tura.objects.filter(
                aktivan=aktivan_filter,
                datum_polaska__year=godina
            ).order_by('datum_polaska')
    else:
        ture = Tura.objects.filter(aktivan=aktivan_filter).order_by('datum_polaska')

    # Ukupne vrijednosti
    total_km = ture.aggregate(Sum('kilometraza'))['kilometraza__sum'] or 0
    total_zaduz = ture.aggregate(Sum('zaduzenje'))['zaduzenje__sum'] or 0
    total_razduz = ture.aggregate(Sum('razduzenje'))['razduzenje__sum'] or 0
    total_razlika = ture.aggregate(Sum('razlika'))['razlika__sum'] or 0
    total_iznos = ture.aggregate(Sum('iznos_ture'))['iznos_ture__sum'] or 0
    total_dnevnice = ture.aggregate(Sum('dnevnice'))['dnevnice__sum'] or 0
    total_cekanje = ture.aggregate(Sum('cekanje'))['cekanje__sum'] or 0
    
    svi_mjeseci = range(1, 13)
    trenutna_godina = datetime.now().year
    
    vozila = Vozilo.objects.all()
    upozorenja = []
    for v in vozila:
        if v.registracija_blizu():
            upozorenja.append(f"⚠️ Vozilu {v.ime} ističe registracija {v.vrijeme_registracije.strftime('%d.%m.%Y')}.")
        if v.servis_blizu():
            upozorenja.append(f"🔧 Vozilo {v.ime} ima servis {v.servis.strftime('%d.%m.%Y')}.")
        
        if v.registracija_istekla():
            upozorenja.append(f"❌ Vozilu {v.ime} je istekla registracija {v.vrijeme_registracije.strftime('%d.%m.%Y')}!")
        if v.servis_istekao():
            upozorenja.append(f"❌ Vozilo {v.ime} je prošao servis {v.servis.strftime('%d.%m.%Y')}!")
        if v.periodicni_pregled_blizu():
            upozorenja.append(f"📋 Vozilu {v.ime} uskoro ističe periodični pregled ({v.periodicni_pregled.strftime('%d.%m.%Y')}).")
        if v.periodicni_pregled_istekao():
            upozorenja.append(f"❌ Vozilu {v.ime} je istekao periodični pregled {v.periodicni_pregled.strftime('%d.%m.%Y')}!")
        if v.bazdarenje_blizu():
            upozorenja.append(f"⏱️ Vozilu {v.ime} uskoro ističe baždarenje tahografa ({v.bazdar_tahografa.strftime('%d.%m.%Y')}).")
        if v.bazdarenje_isteklo():
            upozorenja.append(f"❌ Vozilu {v.ime} je isteklo baždarenje tahografa {v.bazdar_tahografa.strftime('%d.%m.%Y')}!")
        if v.pozarni_aparati_blizu():
            upozorenja.append(f"🔥 Vozilu {v.ime} uskoro ističe pregled požarnih aparata ({v.pozarni_aparati.strftime('%d.%m.%Y')}).")
        if v.pozarni_aparati_istekli():
            upozorenja.append(f"❌ Vozilu {v.ime} su istekli požarni aparati {v.pozarni_aparati.strftime('%d.%m.%Y')}!")
        if v.tu_potvrda_blizu():
            upozorenja.append(f"📄 Vozilu {v.ime} uskoro ističe TU potvrda ({v.tu_potvrda.strftime('%d.%m.%Y')}).")
        if v.tu_potvrda_istekla():
            upozorenja.append(f"❌ Vozilu {v.ime} je istekla TU potvrda {v.tu_potvrda.strftime('%d.%m.%Y')}!")
            
    # Dohvati GET parametre
    # === FILTERI ZA RADNE NALOGE ===
    status_rn = request.GET.get('status_rn', 'aktivni')  # default: aktivni
    aktivan_rn = status_rn != 'zavrseni'

    mjesec_rn = request.GET.get('mjesec_rn')
    godina_rn = request.GET.get('godina_rn')
    tjedan_rn = request.GET.get('tjedan_rn')
    godina_tjedan_rn = request.GET.get('godina_tjedan_rn')

    mjesec_rn_int = parse_int(mjesec_rn)
    godina_rn_int = parse_int(godina_rn)
    tjedan_rn_int = parse_int(tjedan_rn)
    godina_tjedan_rn_int = parse_int(godina_tjedan_rn)

    radni_nalozi = RadniNalog.objects.select_related('tura', 'tura__vozac').filter(aktivan=aktivan_rn)

    # --- Filter po tjednu ---
    if tjedan_rn_int is not None and godina_tjedan_rn_int is not None:
        try:
            start = datetime(godina_tjedan_rn_int, 1, 4)
            start -= timedelta(days=start.weekday())
            start += timedelta(weeks=tjedan_rn_int - 1)
            end = start + timedelta(days=6)
            radni_nalozi = radni_nalozi.filter(
                tura__datum_polaska__date__gte=start.date(),
                tura__datum_polaska__date__lte=end.date()
            )
        except:
            pass

    # 2. MJESEC + GODINA (ili samo GODINA)
    if godina_rn_int is not None:
        if mjesec_rn_int is not None:
            # Mjesec + Godina
            radni_nalozi = radni_nalozi.filter(
                tura__datum_polaska__month=mjesec_rn_int,
                tura__datum_polaska__year=godina_rn_int
        ).order_by('-tura__datum_polaska')
        else:
            # SAMO GODINA
            radni_nalozi = radni_nalozi.filter(
                tura__datum_polaska__year=godina_rn_int
            ).order_by('-tura__datum_polaska')
    else:
        radni_nalozi = radni_nalozi.filter(
                aktivan=aktivan_rn
            ).order_by('-tura__datum_polaska')


    # === Generiraj listu tjedana za trenutnu godinu ===
    trenutna_godina_int = datetime.now().year
    tjedni = []
    prvi_dan = datetime(trenutna_godina_int, 1, 1)
    prvi_pon = prvi_dan - timedelta(days=prvi_dan.weekday())

    for t in range(1, 54):
        pocetak = prvi_pon + timedelta(weeks=t-1)
        kraj = pocetak + timedelta(days=6)
        if pocetak.year > trenutna_godina_int and kraj.year > trenutna_godina_int:
            break
        if pocetak.year == trenutna_godina_int or kraj.year == trenutna_godina_int:
            tjedni.append({
                'broj': t,
                'godina': pocetak.year,
                'label': f"Tjedan {t} – {pocetak.strftime('%d.%m.')} – {kraj.strftime('%d.%m.%Y')}"
            })

    return render(request, 'homepage.html', {
        'ture': ture,
        'total_km': total_km,
        'total_zaduz': total_zaduz,
        'total_razduz': total_razduz,
        'total_razlika': total_razlika,
        'total_iznos': total_iznos,
        'total_dnevnice': total_dnevnice,
        'total_cekanje': total_cekanje,
        'upozorenja': upozorenja,
        'svi_mjeseci': svi_mjeseci,
        'trenutna_godina': trenutna_godina,
        'odabrani_mjesec':int(mjesec) if mjesec else None,
        'odabrana_godina': int(godina) if godina else None,
        'odabrani_status': status,
        'radni_nalozi': radni_nalozi,
        'tjedni': tjedni,
        'odabrani_mjesec_rn': int(mjesec_rn) if mjesec_rn else None,
        'odabrana_godina_rn': int(godina_rn) if godina_rn else None,
        'tjedan_rn': int(tjedan_rn) if tjedan_rn else None,
        'godina_tjedan_rn': int(godina_tjedan_rn) if godina_tjedan_rn else None,
    })
    
@login_required
def zavrsi_radni_nalog(request, rn_id):
    rn = get_object_or_404(RadniNalog, id=rn_id)
    if request.method == 'POST':
        rn.aktivan = False
        rn.save()
        messages.success(request, f"Radni nalog #{rn.id} je završen.")#type: ignore
        return redirect('homepage')
    return render(request, 'radni_nalog/zavrsi.html', {'rn': rn})

@login_required
def unos_ture(request):
    if request.method == 'POST':
        form = TuraForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Tura uspješno dodana.")
            return redirect('unos_ture')
    else:
        form = TuraForm()

    ture = Tura.objects.filter(aktivan=True).order_by('datum_polaska')

    # Ukupne vrijednosti
    total_km = ture.aggregate(Sum('kilometraza'))['kilometraza__sum'] or 0
    total_zaduz = ture.aggregate(Sum('zaduzenje'))['zaduzenje__sum'] or 0
    total_razduz = ture.aggregate(Sum('razduzenje'))['razduzenje__sum'] or 0
    total_razlika = ture.aggregate(Sum('razlika'))['razlika__sum'] or 0
    total_iznos = ture.aggregate(Sum('iznos_ture'))['iznos_ture__sum'] or 0
    total_dnevnice = ture.aggregate(Sum('dnevnice'))['dnevnice__sum'] or 0
    total_cekanje = ture.aggregate(Sum('cekanje'))['cekanje__sum'] or 0

    # Prenos u sljedeći mjesec 
    prenos = total_razlika - total_dnevnice - total_cekanje

    return render(request, 'unos_ture.html', {
        'form': form,
        'ture': ture,
        'total_km': total_km,
        'total_zaduz': total_zaduz,
        'total_razduz': total_razduz,
        'total_razlika': total_razlika,
        'total_iznos': total_iznos,
        'total_dnevnice': total_dnevnice,
        'total_cekanje': total_cekanje,
        'prenos': prenos,
    })

@login_required
def unos_vozaca(request):
    if request.method == 'POST':
        form = VozacForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('popis_vozaca')
    else:
        form = VozacForm()

    vozac = Vozac.objects.all()
    return render(request, 'popis_vozaca.html', {'form': form, 'vozaci': vozac})

@login_required
def profil_vozaca(request, vozac_id):
    vozac = get_object_or_404(Vozac, id=vozac_id)
    
    mjesec = parse_int(request.GET.get('mjesec'), datetime.now().month) if request.GET.get('mjesec') else None
    godina = parse_int(request.GET.get('godina'), datetime.now().year)
    
    # === NOVO: TJEDAN FILTER ===
    iso = datetime.now().isocalendar()
    
    tjedan_rn_str = request.GET.get('tjedan_rn', '').strip()
    godina_tjedan_rn_str = request.GET.get('godina_tjedan_rn', '').strip()

    # Ako je parametar prazan ili 'None' → ne filtriramo po tjednu
    tjedan_rn = None
    godina_tjedan_rn = None

    if tjedan_rn_str and tjedan_rn_str.lower() != 'none':
        tjedan_rn = parse_int(tjedan_rn_str)

    if godina_tjedan_rn_str and godina_tjedan_rn_str.lower() != 'none':
        godina_tjedan_rn = parse_int(godina_tjedan_rn_str)

    # Ako nema godine tjedna, a ima tjedan → fallback na tekuću ISO godinu
    if tjedan_rn is not None and godina_tjedan_rn is None:
        godina_tjedan_rn = iso.year
    
    if mjesec and godina:
        ture = Tura.objects.filter(
            vozac=vozac,
            datum_polaska__month=mjesec,
            datum_polaska__year=godina
            ).order_by('datum_polaska')
    elif godina:
        ture = Tura.objects.filter(
            vozac=vozac,
            datum_polaska__year=godina
            ).order_by('datum_polaska')
    else:
        ture = Tura.objects.filter(vozac=vozac).order_by('datum_polaska')
        
    radni_nalozi_vozaca = RadniNalog.objects.filter(
        tura__vozac=vozac,
        tura__datum_polaska__year=godina_tjedan_rn if godina_tjedan_rn else None
    ).select_related('tura').order_by('-tura__datum_polaska')

    # Primjenjujemo tjedan filter SAMO ako imamo oba parametra
    if tjedan_rn is not None and godina_tjedan_rn is not None:
        try:
            start = datetime(godina_tjedan_rn, 1, 4)
            start -= timedelta(days=start.weekday())          # ponedjeljak
            start_date = start + timedelta(weeks=tjedan_rn - 1)
            end_date = start_date + timedelta(days=6)

            radni_nalozi_vozaca = radni_nalozi_vozaca.filter(
                tura__datum_polaska__date__gte=start_date.date(),
                tura__datum_polaska__date__lte=end_date.date()
            )
        except (ValueError, OverflowError):
            # Neispravan datum → ne filtriramo (ili možeš dodati poruku)
            pass 
        
    if request.method == 'POST':
        form = VozacUpdateForm(request.POST, instance=vozac)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Izmjene su uspješno spremljene.")
            for tura in ture:
                if tura.iznos_ture:
                    tura.dnevnice = round(tura.iznos_ture * vozac.postotak , 2)
                    tura.save(update_fields=['dnevnice'])
            return redirect('profil_vozaca', vozac_id=vozac.id) # type: ignore
    else:
        form = VozacUpdateForm(instance=vozac)

    # Izračun suma
    total_km = ture.aggregate(Sum('kilometraza'))['kilometraza__sum'] or 0
    total_zaduz = ture.aggregate(Sum('zaduzenje'))['zaduzenje__sum'] or 0
    total_razduz = ture.aggregate(Sum('razduzenje'))['razduzenje__sum'] or 0
    total_razlika = ture.aggregate(Sum('razlika'))['razlika__sum'] or 0
    total_iznos = ture.aggregate(Sum('iznos_ture'))['iznos_ture__sum'] or 0
    total_dnevnice = round(ture.aggregate(Sum('dnevnice'))['dnevnice__sum'] or 0, 2)
    total_cekanje = ture.aggregate(Sum('cekanje'))['cekanje__sum'] or 0

    # Prenos i bilanca
    bilanca = round(( total_razlika - total_dnevnice - total_cekanje + vozac.zaduzenje_prethodni_mjesec + vozac.uplaceno_na_banku),2)
    
    svi_mjeseci = range(1, 13)
    trenutna_godina = datetime.now().year
    
    # Generiraj listu tjedana (isto kao na homepage-u)
    tjedni = []
    prvi_dan = datetime(trenutna_godina, 1, 1)
    prvi_pon = prvi_dan - timedelta(days=prvi_dan.weekday())
    for t in range(1, 54):
        pocetak = prvi_pon + timedelta(weeks=t-1)
        kraj = pocetak + timedelta(days=6)
        if pocetak.year >= trenutna_godina - 1 and pocetak.year <= trenutna_godina + 1:
            if pocetak.year == trenutna_godina or kraj.year == trenutna_godina:
                tjedni.append({
                    'broj': t,
                    'godina': pocetak.year,
                    'label': f"Tjedan {t} – {pocetak.strftime('%d.%m.')} - {kraj.strftime('%d.%m.%Y')}"
                })

    return render(request, 'profil_vozaca.html', {
        'vozac': vozac,
        'ture': ture,
        'total_km': total_km,
        'total_zaduz': total_zaduz,
        'total_razduz': total_razduz,
        'total_razlika': total_razlika,
        'total_iznos': total_iznos,
        'total_dnevnice': total_dnevnice,
        'total_cekanje': total_cekanje,
        'bilanca': bilanca,
        'form': form,
        'svi_mjeseci': svi_mjeseci,
        'trenutna_godina': trenutna_godina,
        'odabrani_mjesec': int(mjesec) if mjesec else None,
        'odabrana_godina': int(godina) if godina else None,
        'radni_nalozi_vozaca': radni_nalozi_vozaca,
        'tjedni': tjedni,
        'tjedan_rn': tjedan_rn,
        'godina_tjedan_rn': godina_tjedan_rn or trenutna_godina,
    })

@login_required    
def dodavanje_vozaca(request):
    if request.method == 'POST':
        form = VozacForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('popis_vozaca')
    else:
        form = VozacForm()

    return render(request, 'dodavanje_vozaca.html', {'form': form})

@login_required
def zavrsi_turu(request, tura_id):
    tura = get_object_or_404(Tura, id=tura_id)
    tura.aktivan = False
    tura.save()
    return redirect('homepage')  # Vrati korisnika na popis aktivnih tura

@login_required
def profil_ture(request, tura_id):
    tura = get_object_or_404(Tura, id=tura_id)
    
    if request.method == 'POST':
        form = TuraForm(request.POST, instance=tura)
        if form.is_valid():
            form.save()
            osvjezi_radni_nalog(tura=tura)
            messages.success(request, "✅ Izmjene su uspješno spremljene.")
            return redirect('profil_ture', tura_id=tura.id) # type: ignore
    else:
        form = TuraForm(instance=tura)

    return render(request, 'profil_ture.html', {'tura': tura, 'form': form})

@login_required
def popis_vozila(request):
    vozila = Vozilo.objects.all().order_by('ime')
    return render(request, 'vozila/popis_vozila.html', {'vozila': vozila})

@login_required
def detalji_vozila(request, vozilo_id):
    vozilo = get_object_or_404(Vozilo, id=vozilo_id)
    naputci = vozilo.naputci.all().order_by('-datum') # type: ignore
    trenutna_godina = datetime.now().year
    godine = range(trenutna_godina - 5, trenutna_godina + 1)

    if request.method == 'POST':
        form = NaputakForm(request.POST)
        if form.is_valid():
            naputak = form.save(commit=False)
            naputak.vozilo = vozilo
            naputak.save()
            return redirect('detalji_vozila', vozilo_id=vozilo.id)# type: ignore
    else:
        form = NaputakForm()

    return render(request, 'vozila/detalji_vozila.html', {
        'vozilo': vozilo,
        'naputci': naputci,
        'form': form,
        'godine': godine, 
        'trenutna_godina': trenutna_godina
    })

@login_required
def dodaj_vozilo(request):
    if request.method == 'POST':
        form = VoziloForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('popis_vozila')
    else:
        form = VoziloForm()
    return render(request, 'vozila/dodaj_vozilo.html', {'form': form})

@login_required
def uredi_vozilo(request, vozilo_id):
    vozilo = get_object_or_404(Vozilo, id=vozilo_id)
    if request.method == 'POST':
        form = VoziloForm(request.POST, instance=vozilo)
        if form.is_valid():
            form.save()
            return redirect('detalji_vozila', vozilo_id=vozilo.id) # type: ignore
    else:
        form = VoziloForm(instance=vozilo)
    return render(request, 'vozila/uredi_vozilo.html', {'form': form, 'vozilo': vozilo})

@login_required
def obrisi_vozilo(request, vozilo_id):
    vozilo = get_object_or_404(Vozilo, id=vozilo_id)
    if request.method == 'POST':
        vozilo.delete()
        return redirect('popis_vozila')
    return render(request, 'vozila/obrisi_vozilo.html', {'vozilo': vozilo})

@login_required
def uredi_naputak(request, naputak_id):
    naputak = get_object_or_404(Naputak, id=naputak_id)
    vozilo = naputak.vozilo

    if request.method == 'POST':
        form = NaputakForm(request.POST, instance=naputak)
        if form.is_valid():
            form.save()
            return redirect('detalji_vozila', vozilo_id=vozilo.id)
    else:
        form = NaputakForm(instance=naputak)

    return render(request, 'uredi_naputak.html', {'form': form, 'naputak': naputak, 'vozilo': vozilo})

@login_required
def obrisi_naputak(request, naputak_id):
    naputak = get_object_or_404(Naputak, id=naputak_id)
    vozilo = naputak.vozilo

    if request.method == 'POST':
        naputak.delete()
        return redirect('detalji_vozila', vozilo_id=vozilo.id)# type: ignore

    return render(request, 'obrisi_naputak.html', {'naputak': naputak, 'vozilo': vozilo})

@login_required
def export_vozac_pdf(request, vozac_id):
    vozac = get_object_or_404(Vozac, id=vozac_id)

    mjesec = request.GET.get('mjesec')
    godina = request.GET.get('godina')

    if mjesec and godina:
        ture = Tura.objects.filter(vozac=vozac, datum_polaska__month=mjesec, datum_polaska__year=godina).order_by('datum_polaska')
        naziv_perioda = f"{int(mjesec)}. mjesec {godina}"
    elif godina:
        ture = Tura.objects.filter(vozac=vozac, datum_polaska__year=godina).order_by('datum_polaska')
        naziv_perioda = f"{godina}. godina"
    else:
        ture = Tura.objects.filter(vozac=vozac).order_by('datum_polaska')
        naziv_perioda = datetime.now().strftime("%m.%Y")

    if not ture.exists():
        messages.info(request, "Nema tura za odabrani period.")
        return redirect('profil_vozaca', vozac_id=vozac.id)#type: ignore

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Ture_{vozac.ime}_{naziv_perioda.replace(" ", "_")}.pdf"'

    font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'Arial.ttf')
    pdfmetrics.registerFont(TTFont('Arial', font_path))

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=20, bottomMargin=30)#type: ignore
    elements = []
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CustomTitle', fontName='Arial', fontSize=20, alignment=1, spaceAfter=10))
    styles.add(ParagraphStyle(name='CustomNormal', fontName='Arial', fontSize=10, leading=12))
    styles.add(ParagraphStyle(name='BilancaLabel', fontName='Arial', fontSize=12))
    styles.add(ParagraphStyle(name='BilancaIznos', fontName='Arial', fontSize=14))
    
    # Odabir loga – default je ihlogistika
    logo_choice = request.GET.get('logo', 'ihlogistika')

    # Definiraj putanje do loga
    if logo_choice == 'ihtransport':
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo1.jpg')
    else:
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo2.jpg')


    elements.append(Spacer(1, 8))

    if os.path.exists(logo_path):
        try:
            logo_img = Image(logo_path, width=120, height=60)  # prilagodi dimenzije po potrebi
            logo_img.hAlign = 'LEFT'
            elements.append(logo_img)
            elements.append(Spacer(1, 12))
        except Exception:
            pass
    else:
        elements.append(Spacer(1, 20))

    # Naslov
    elements.append(Paragraph(f"TURE – {vozac.ime}", styles['CustomTitle']))
    elements.append(Paragraph(naziv_perioda, ParagraphStyle(name='Sub', fontName='Arial', fontSize=14, alignment=1, spaceAfter=25)))

    # Tablica
    data = [["Relacija", "Polazak", "Povratak", "Km", "Zaduženje", "Razduženje", "Razlika", "Iznos", "Dnevnice", "Čekanje"]]
    ukupno_km = ukupno_zaduz = ukupno_razduz = ukupno_razlika = ukupno_iznos = ukupno_dnevnice = ukupno_cekanje = 0

    for t in ture:
        data.append([
            Paragraph(t.relacija or "", styles['CustomNormal']),
            Paragraph(t.datum_polaska.strftime('%d.%m.%Y') if t.datum_polaska else "", styles['CustomNormal']),
            Paragraph(t.datum_dolaska.strftime('%d.%m.%Y') if t.datum_dolaska else "", styles['CustomNormal']),
            Paragraph(str(t.kilometraza) if t.kilometraza else "", styles['CustomNormal']),
            Paragraph(f"{t.zaduzenje:.2f}" if t.zaduzenje is not None else "", styles['CustomNormal']),
            Paragraph(f"{t.razduzenje:.2f}" if t.razduzenje is not None else "", styles['CustomNormal']),
            Paragraph(f"{t.razlika:.2f}" if t.razlika is not None else "", styles['CustomNormal']),
            Paragraph(f"{t.iznos_ture:.2f}" if t.iznos_ture is not None else "", styles['CustomNormal']),
            Paragraph(f"{t.dnevnice:.2f}" if t.dnevnice is not None else "", styles['CustomNormal']),
            Paragraph(f"{t.cekanje:.2f}" if t.cekanje is not None else "", styles['CustomNormal']),
        ])#type: ignore
        ukupno_km += t.kilometraza or 0
        ukupno_zaduz += t.zaduzenje or 0
        ukupno_razduz += t.razduzenje or 0
        ukupno_razlika += t.razlika or 0
        ukupno_iznos += t.iznos_ture or 0
        ukupno_dnevnice += t.dnevnice or 0
        ukupno_cekanje += t.cekanje or 0

    data.append(["", "", "",
                 Paragraph(f"<b>{ukupno_km}</b>", styles['CustomNormal']),
                 Paragraph(f"<b>{ukupno_zaduz:.2f}</b>", styles['CustomNormal']),
                 Paragraph(f"<b>{ukupno_razduz:.2f}</b>", styles['CustomNormal']),
                 Paragraph(f"<b>{ukupno_razlika:.2f}</b>", styles['CustomNormal']),
                 Paragraph(f"<b>{ukupno_iznos:.2f}</b>", styles['CustomNormal']),
                 Paragraph(f"<b>{ukupno_dnevnice:.2f}</b>", styles['CustomNormal']),
                 Paragraph(f"<b>{ukupno_cekanje:.2f}</b>", styles['CustomNormal'])])#type: ignore

    table = Table(data, colWidths=[140, 70, 70, 60, 70, 70, 70, 70, 70, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#dddddd')),
        ('GRID', (0,0), (-1,-1), 0.8, colors.black),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#bbbbbb')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (3,1), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (-1,0), 'Arial'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(table)

    # Bilanca dolje desno
    bilanca = round(ukupno_razlika - ukupno_dnevnice - ukupno_cekanje + vozac.zaduzenje_prethodni_mjesec + vozac.uplaceno_na_banku, 2)
    elements.append(Spacer(1, 20))
    bilanca_table = Table([[Paragraph("Bilanca:", styles['BilancaLabel']),
                            Paragraph(f"<b>{bilanca:,.2f} {vozac.valuta}</b>".replace(',', 'X').replace('.', ',').replace('X', '.'), styles['BilancaIznos'])]],
                          colWidths=[100, 100])
    bilanca_table.hAlign = 'RIGHT'
    elements.append(bilanca_table)

    doc.build(elements)
    return response

@login_required
def dodaj_radni_nalog(request):
    if request.method == 'POST':
        form = RadniNalogForm(request.POST)
        if form.is_valid():
            radni_nalog = form.save()
            messages.success(request, "✅ Radni nalog uspješno kreiran.")
            return redirect('radni_nalog_detail', radni_nalog_id=radni_nalog.id)
    else:
        form = RadniNalogForm()
    return render(request, 'radni_nalog/dodaj.html', {'form': form})


@login_required
def radni_nalog_detail(request, radni_nalog_id):
    radni_nalog = get_object_or_404(RadniNalog, id=radni_nalog_id)
    return render(request, 'radni_nalog/detail.html', {'radni_nalog': radni_nalog})


@login_required
def uredi_radni_nalog(request, radni_nalog_id):
    radni_nalog = get_object_or_404(RadniNalog, id=radni_nalog_id)
    if request.method == 'POST':
        form = RadniNalogForm(request.POST, instance=radni_nalog)
        if form.is_valid():
            form.save()  # save() poziva izracun_dnevnica() preko modela
            messages.success(request, "Radni nalog ažuriran, dnevnice preračunate.")
            return redirect('radni_nalog_detail', radni_nalog_id=radni_nalog.id)#type: ignore
    else:
        form = RadniNalogForm(instance=radni_nalog)
    
    return render(request, 'radni_nalog/uredi.html', {
        'form': form,
        'radni_nalog': radni_nalog,
    })
    

@login_required
def cijene_dnevnica(request):
    # Dohvati sve cijene
    cijene = CijenaDnevnica.objects.all().order_by('drzava')

    if request.method == 'POST':
        for cijena in cijene:
            field_name = f'iznos_{cijena.id}'#type: ignore
            novi_iznos_str = request.POST.get(field_name, '').strip()

            if novi_iznos_str:
                try:
                    # Podrška za zarez i točku: 90,50 ili 90.50 → 90.50
                    novi_iznos = float(novi_iznos_str.replace(',', '.'))
                    if novi_iznos < 0:
                        raise ValueError
                    cijena.iznos = novi_iznos
                    cijena.save()
                except ValueError:
                    messages.error(request, f"Neispravan iznos za {cijena.get_drzava_display()}: '{novi_iznos_str}'")#type: ignore
                    continue

        messages.success(request, "Cijene dnevnica su uspješno ažurirane.")
        return redirect('cijene_dnevnica')

    # --- PRIKAZ: Formatiraj iznos sa točkom i 2 decimale (za input) ---
    for c in cijene:
        c.iznos_input = f"{float(c.iznos):.2f}"  #type: ignore
        
    radni_nalozi = RadniNalog.objects.all() 
    
    for rn in radni_nalozi:
        rn.izracun_dnevnica()
        rn.save()

    return render(request, 'cijene_dnevnica.html', {
        'cijene': cijene
    })

# @login_required
# def export_vozacev_tjedan_pdf(request, vozac_id):
#     vozac = get_object_or_404(Vozac, id=vozac_id)
#     tjedan = parse_int(request.GET.get('tjedan_rn'))
#     godina_tjedan = parse_int(request.GET.get('godina_tjedan_rn'))

#     if not tjedan or not godina_tjedan:
#         messages.error(request, "Odaberi tjedan za eksport!")
#         return redirect('profil_vozaca', vozac_id=vozac.id)#type: ignore

#     try:
#         start = datetime(godina_tjedan, 1, 4)
#         start -= timedelta(days=start.weekday())
#         start_date = start + timedelta(weeks=tjedan - 1)
#         end_date = start_date + timedelta(days=6)
#     except:
#         messages.error(request, "Neispravan tjedan ili godina!")
#         return redirect('profil_vozaca', vozac_id=vozac.id)#type: ignore

#     radni_nalozi = RadniNalog.objects.filter(
#         tura__vozac=vozac,
#         tura__datum_polaska__date__gte=start_date.date(),
#         tura__datum_polaska__date__lte=end_date.date()
#     ).select_related('tura').order_by('tura__datum_polaska')

#     if not radni_nalozi.exists():
#         messages.info(request, f"Nema radnih naloga za tjedan {tjedan}.")
#         return redirect('profil_vozaca', vozac_id=vozac.id)#type: ignore

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = f'attachment; filename="RN_{vozac.ime}_Tjedan_{tjedan}_{godina_tjedan}.pdf"'

#     font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'Arial.ttf')
#     pdfmetrics.registerFont(TTFont('Arial', font_path))

#     doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=20, bottomMargin=30)#type: ignore
#     elements = []

#     # === DVA LOGA – GORNJI LIJEVI KUT ===
#     logo1_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo1.jpg')
#     logo2_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo2.jpg')

#     logos = []
#     if os.path.exists(logo1_path):
#         img1 = Image(logo1_path, width=75, height=48)
#         img1.hAlign = 'LEFT'
#         logos.append(img1)
#     if os.path.exists(logo2_path):
#         img2 = Image(logo2_path, width=75, height=48)
#         img2.hAlign = 'LEFT'
#         logos.append(img2)

#     if logos:
#         if len(logos) == 2:
#             logo_table = Table([[logos[0], logos[1]]], colWidths=[85, 85])
#         else:
#             logo_table = Table([[logos[0]]], colWidths=[85])
        
#         logo_table.setStyle(TableStyle([
#             ('ALIGN', (0,0), (-1,-1), 'LEFT'),
#             ('VALIGN', (0,0), (-1,-1), 'TOP'),
#             ('LEFTPADDING', (0,0), (-1,-1), 0),
#             ('RIGHTPADDING', (0,0), (-1,-1), 40),
#             ('BOTTOMPADDING', (0,0), (-1,-1), 0),
#         ]))
#         elements.append(KeepInFrame(600, 100, [logo_table], hAlign='LEFT', vAlign='TOP'))
#         elements.append(Spacer(1, 12))

#     styles = getSampleStyleSheet()
#     styles.add(ParagraphStyle(name='CustomTitle', fontName='Arial', fontSize=20, alignment=1, spaceAfter=10))
#     styles.add(ParagraphStyle(name='CustomNormal', fontName='Arial', fontSize=10, leading=12))

#     elements.append(Paragraph(f"RADNI NALOZI – {vozac.ime}", styles['CustomTitle']))
#     elements.append(Paragraph(f"Tjedan {tjedan} / {godina_tjedan}  •  {start_date.strftime('%d.%m.')} – {end_date.strftime('%d.%m.%Y')}",
#                               ParagraphStyle(name='Sub', fontName='Arial', fontSize=14, alignment=1, spaceAfter=25)))

#     # Tablica
#     data = [["Relacija", "Polazak", "Povratak", "Država", f"Tuzemne ({vozac.valuta})", f"Inozemne ({vozac.valuta})", f"Ukupno ({vozac.valuta})"]]
#     ukupno_tuzemne = ukupno_inozemne = ukupno_sve = 0

#     for rn in radni_nalozi:
#         tura = rn.tura
#         ukupno_po_nalogu = rn.tuzemne_dnevnice + rn.inozemne_dnevnice#type: ignore
#         data.append([
#             Paragraph(tura.relacija or "", styles['CustomNormal']),
#             Paragraph(tura.datum_polaska.strftime('%d.%m.') if tura.datum_polaska else "", styles['CustomNormal']),
#             Paragraph(tura.datum_dolaska.strftime('%d.%m.') if tura.datum_dolaska else "", styles['CustomNormal']),
#             Paragraph(rn.konacna_drzava or "", styles['CustomNormal']),
#             Paragraph(f"{rn.tuzemne_dnevnice:.2f}", styles['CustomNormal']),
#             Paragraph(f"{rn.inozemne_dnevnice:.2f}", styles['CustomNormal']),
#             Paragraph(f"<b>{ukupno_po_nalogu:.2f}</b>", styles['CustomNormal']),
#         ])#type: ignore
#         ukupno_tuzemne += rn.tuzemne_dnevnice#type: ignore
#         ukupno_inozemne += rn.inozemne_dnevnice#type: ignore
#         ukupno_sve += ukupno_po_nalogu

#     data.append([
#         Paragraph("<b>UKUPNO ZA TJEDAN</b>", styles['CustomNormal']), "", "", "",
#         Paragraph(f"<b>{ukupno_tuzemne:.2f}</b>", styles['CustomNormal']),
#         Paragraph(f"<b>{ukupno_inozemne:.2f}</b>", styles['CustomNormal']),
#         Paragraph(f"<b>{ukupno_sve:.2f} {vozac.valuta}</b>", styles['CustomNormal']),
#     ])#type: ignore

#     table = Table(data, colWidths=[165, 70, 70, 100, 80, 80, 95])
#     table.setStyle(TableStyle([
#         ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#dddddd')),
#         ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#bbbbbb')),
#         ('GRID', (0,0), (-1,-1), 0.8, colors.black),
#         ('ALIGN', (0,0), (-1,-1), 'CENTER'),
#         ('ALIGN', (4,1), (-1,-1), 'RIGHT'),
#         ('FONTSIZE', (0,0), (-1,-1), 10),
#         ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
#         ('LEFTPADDING', (0,0), (-1,-1), 6),
#         ('RIGHTPADDING', (0,0), (-1,-1), 6),
#         ('TOPPADDING', (0,0), (-1,-1), 8),
#         ('BOTTOMPADDING', (0,0), (-1,-1), 8),
#     ]))
#     elements.append(table)

#     doc.build(elements)
#     return response

@login_required
def export_naputci_pdf(request, vozilo_id):
    vozilo = get_object_or_404(Vozilo, id=vozilo_id)
    godina = int(request.GET.get('godina', datetime.now().year))
    
    # ─── Odabir loga ────────────────────────────────────────────────
    logo_choice = request.GET.get('logo', 'ihlogistika')  # default: ihlogistika

    if logo_choice == 'ihtransport':
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo1.jpg')
    else:
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo2.jpg')

    
    # Filtriraj naputke za vozilo i godinu
    naputci = Naputak.objects.filter(vozilo=vozilo, datum__year=godina).order_by('datum')
    
    # Postavke PDF-a
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="naputci_{vozilo.ime}_{godina}.pdf"'
    
    # Kreiraj PDF dokument (landscape za širu tablicu)
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    # Registriraj font (za hrvatske znakove, preuzeto iz vašeg postojećeg koda)
    font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'Arial.ttf')  
    pdfmetrics.registerFont(TTFont('Arial', font_path))
    
    # Stilovi (kopirano iz vašeg export_vozacev_tjedan_pdf ili slično)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CustomNormal', fontName='Arial', fontSize=10, leading=12))
    styles.add(ParagraphStyle(name='CustomHeading', fontName='Arial', fontSize=14, leading=16, alignment=1))  # Centrirano
    
    elements = []
    
    if logo_path:
        try:
            logo_img = Image(logo_path, width=120, height=60)  # prilagodi dimenzije
            logo_img.hAlign = 'LEFT'
            elements.append(logo_img)
            elements.append(Spacer(1, 12))
        except Exception as e:
            # Ako slika ne može biti učitana – preskoči ili logiraj
            pass
    
    # Naslov
    elements.append(Paragraph(f"Naputci za vozilo: {vozilo.ime} u {godina}. godini", styles['CustomHeading']))
    elements.append(Spacer(1, 12))
    
    # Podaci za tablicu
    data = [['Datum', 'Sadržaj']]  # Header
    
    if not naputci.exists():
        data.append(['Nema naputaka za ovu godinu.', ''])
    else:
        for naputak in naputci:
            data.append([
                naputak.datum.strftime('%d.%m.%Y %H:%M'),
                Paragraph(naputak.sadrzaj, styles['CustomNormal'])  # Omogućuje wrap teksta
            ])
    
    # Tablica
    table = Table(data, colWidths=[150, 500])  # Šire za sadržaj
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dddddd')),
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Arial'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    elements.append(table)
    
    # Footer (opcionalno)
    elements.append(Spacer(1, 24))
    elements.append(Paragraph(f"Generirano: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['CustomNormal']))
    
    doc.build(elements)
    return response

@login_required
def export_vozacev_tjedan_pdf(request, vozac_id):
    vozac = get_object_or_404(Vozac, id=vozac_id)
    vozilo = vozac.vozila.first()
    tjedan = parse_int(request.GET.get('tjedan_rn'))
    godina_tjedan = parse_int(request.GET.get('godina_tjedan_rn'))

    if not tjedan or not godina_tjedan:
        messages.error(request, "Odaberi tjedan za eksport!")
        return redirect('profil_vozaca', vozac_id=vozac.id)#type: ignore

    try:
        start = datetime(godina_tjedan, 1, 4)
        start -= timedelta(days=start.weekday())
        start_date = start + timedelta(weeks=tjedan - 1)
        end_date = start_date + timedelta(days=6)
    except:
        messages.error(request, "Neispravan tjedan ili godina!")
        return redirect('profil_vozaca', vozac_id=vozac.id)#type: ignore

    radni_nalozi = RadniNalog.objects.filter(
        tura__vozac=vozac,
        tura__datum_polaska__date__gte=start_date.date(),
        tura__datum_polaska__date__lte=end_date.date()
    ).select_related('tura').order_by('tura__datum_polaska')

    if not radni_nalozi.exists():
        messages.info(request, f"Nema radnih naloga za tjedan {tjedan}.")
        return redirect('profil_vozaca', vozac_id=vozac.id)#type: ignore

    template_path = os.path.join(BASE_DIR, "excel_template", "template.xlsx")

    for rn in radni_nalozi:
    # 1. Otvori template
        wb = load_workbook(template_path)
        ws = wb["PutniNalog"]
        
        delta = rn.tura.datum_dolaska - rn.tura.datum_polaska

        ukupni_dani = delta.days
        ostatak_sati = delta.seconds // 3600
        
        # 2. Popuni fiksne lokacije
        ws["G4"] = rn.tura.datum_polaska.strftime('%d.%m.%Y') if rn.tura.datum_polaska else ""
        ws["G5"] = rn.tura.datum_dolaska.strftime('%d.%m.%Y') if rn.tura.datum_dolaska else ""
        ws["T4"] = rn.tura.datum_polaska.strftime('%H:%M') if rn.tura.datum_polaska else ""
        ws["T5"] = rn.tura.datum_dolaska.strftime('%H:%M') if rn.tura.datum_dolaska else ""
        ws["O6"] = ukupni_dani
        ws["U6"] = ostatak_sati
        ws["A9"] = "BiH"
        ws["A10"] = rn.konacna_drzava
        ws["L9"] = 1
        ws["L10"] = 1
        ws["O8"] = f"IZNOS {rn.tura.vozac.valuta}"
        ws["T8"] = f"UKUPNO {rn.tura.vozac.valuta}"
        ws["O9"] = rn.tuzemne_dnevnice
        ws["O10"] = rn.inozemne_dnevnice
        ws["T9"] = rn.tuzemne_dnevnice
        ws["T10"] = rn.inozemne_dnevnice
        ws["T14"] = rn.papiri
        ws["T15"] = rn.terminali
        ws["T16"] = rn.cestarine
        ws["AF4"] = rn.tura.broj_putnog_naloga
        ws["AF5"] = datetime.now().strftime('%d.%m.%Y')
        ws["AF16"] = rn.tura.vozac.ime
        ws["AK17"] = "vozač"
        ws["AN18"] = f"BiH - {rn.konacna_drzava}"
        ws["AF19"] = "obavljanja prijevoza"
        ws["AN21"] = rn.tura.datum_polaska.strftime('%d.%m.%Y') if rn.tura.datum_polaska else ""
        ws["AT21"] = rn.tura.datum_dolaska.strftime('%d.%m.%Y') if rn.tura.datum_dolaska else ""
        ws["AK25"] = vozilo.ime   
        ws["AK26"] = rn.tura.relacija    
        
        output_dir = Path.home() / "Desktop" / "Izvještaji_radnih_naloga"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # 3. Kreiraj output ime
        filename = f"Izvještaj_{rn.tura.vozac}_tjedan_{tjedan}_{godina_tjedan}_{rn.id}.xlsx"
        output_path = os.path.join(output_dir, filename)

        # 4. Save new file
        wb.save(output_path)
        messages.success(request, f"Excel generiran i spremljen u: {output_path}")
    return redirect('profil_vozaca', vozac_id=vozac.id)
